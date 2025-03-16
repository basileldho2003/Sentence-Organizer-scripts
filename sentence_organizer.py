import openpyxl
from openpyxl.styles import Font


def classify_sentence_length(sentence):
    """
    Classifies a sentence based on the number of words.
    """
    word_count = len(sentence.split())

    if 1 <= word_count <= 4:
        return "Very Short (1-4 words)"
    elif 5 <= word_count <= 8:
        return "Short (5-8 words)"
    elif 9 <= word_count <= 11:
        return "Medium (9-11 words)"
    elif 12 <= word_count <= 15:
        return "Long (12-15 words)"
    elif word_count >= 16:
        return "Very Long (16+ words)"
    return "Unknown"


def process_sentences(input_file, output_file):
    """
    Reads an Excel file, classifies sentences by length, and stores them in respective category sheets
    with detailed length categories as columns. Headers are formatted in bold.
    """
    # Load input Excel file
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active  # Assume first sheet has sentences

    # Define categories as per columns
    categories = {
        "A": "Agriculture",
        "B": "Culture",
        "C": "Education",
        "D": "News",
        "E": "Business",
        "F": "Healthcare",
        "G": "Sports",
        "H": "Law",
        "I": "Governance",
        "J": "Tourism",
        "K": "Banking and Finance",
    }

    length_categories = [
        "Very Short (1-4 words)",
        "Short (5-8 words)",
        "Medium (9-11 words)",
        "Long (12-15 words)",
        "Very Long (16+ words)",
    ]

    # Dictionary to store categorized sentences
    categorized_sentences = {
        category: {length: [] for length in length_categories}
        for category in categories.values()
    }

    # Process each column
    for col_idx, (col_letter, category_name) in enumerate(categories.items(), start=1):
        for row in sheet.iter_rows(
            min_row=1, min_col=col_idx, max_col=col_idx, values_only=True
        ):
            sentence = row[0]  # Extract sentence
            if isinstance(sentence, str) and sentence.strip():  # Ensure non-empty text
                length_category = classify_sentence_length(sentence)
                categorized_sentences[category_name][length_category].append(sentence)

    # Create a new Excel workbook for categorized data
    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)  # Remove default sheet

    # Write categorized sentences into separate sheets
    for category, length_data in categorized_sentences.items():
        ws = output_wb.create_sheet(title=category)

        # Write headers in the first row (length categories as columns) and make them bold
        for col_idx, header in enumerate(length_categories, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)  # Apply bold formatting

        # Determine the maximum number of sentences per category
        max_rows = max(len(sentences) for sentences in length_data.values())

        # Write sentences into respective columns
        for row_idx in range(max_rows):
            row_data = []
            for length_label in length_categories:
                if row_idx < len(length_data[length_label]):
                    row_data.append(length_data[length_label][row_idx])
                else:
                    row_data.append("")  # Empty cell if no sentence available
            ws.append(row_data)

    # Save the processed workbook
    output_wb.save(output_file)
    print(f"Processing complete. Categorized sentences saved to {output_file}")


# Example usage
input_file_path = "Sentence_dataset_deduplicated_normalized.xlsx"  # Input dataset file
output_file_path = "Categorized_Sentences.xlsx"  # Output file

process_sentences(input_file_path, output_file_path)
