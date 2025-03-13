import openpyxl

# Import the normalization class
from string_normalizer import TextProcessor


def read_excel(file_path):
    """Reads sentences from an Excel file while preserving the column structure."""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        cleaned_row = [cell.strip() if isinstance(cell, str) else None for cell in row]
        data.append(cleaned_row)

    return data


def remove_duplicates_and_shift_up(data):
    """Removes duplicate sentences while shifting non-empty cells upward in each column."""
    seen = set()
    num_cols = len(data[0])
    columns = [[] for _ in range(num_cols)]

    for row in data:
        for col_idx, cell in enumerate(row):
            if cell and cell not in seen:
                seen.add(cell)
                columns[col_idx].append(cell)

    max_rows = max(len(col) for col in columns)
    deduplicated_data = [
        [
            columns[col_idx][row_idx] if row_idx < len(columns[col_idx]) else None
            for col_idx in range(num_cols)
        ]
        for row_idx in range(max_rows)
    ]

    return deduplicated_data


def normalize_text_in_excel(data):
    """Applies text normalization (e.g., converting numbers to words) to all cells."""
    processor = TextProcessor()

    normalized_data = []
    for row in data:
        normalized_row = [
            processor.process_text(cell) if cell else None for cell in row
        ]
        normalized_data.append(normalized_row)

    return normalized_data


def write_excel(file_path, data):
    """Writes deduplicated and normalized sentences back to an Excel file."""
    wb = openpyxl.Workbook()
    sheet = wb.active

    for row in data:
        sheet.append(row)

    wb.save(file_path)


# Main execution
input_file = "Sentence_dataset.xlsx"
output_file = "Sentence_dataset_deduplicated_normalized.xlsx"

data = read_excel(input_file)
deduplicated_data = remove_duplicates_and_shift_up(data)
normalized_data = normalize_text_in_excel(deduplicated_data)
write_excel(output_file, normalized_data)

print(f"Deduplicated and normalized data saved to: {output_file}")
